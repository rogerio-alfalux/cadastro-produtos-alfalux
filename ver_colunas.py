#!/usr/bin/env python3
import openpyxl
wb = openpyxl.load_workbook('/home/ubuntu/upload/TabelavendasLuminárias25.06.26final.xlsx', data_only=True)
ws = wb.active
print(f"Max col: {ws.max_column}")
print()
# Mostrar cabeçalho completo
print("=== LINHA 1 (CABEÇALHO) ===")
for col in range(1, ws.max_column + 1):
    val = ws.cell(1, col).value
    if val:
        print(f"  Col {col}: {val}")

print()
print("=== LINHAS 18-50 (MOON + próximos) - TODAS AS COLUNAS ===")
for r in range(18, 51):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(r, col).value
        if v is not None:
            vals.append(f"C{col}={repr(str(v)[:25])}")
    if vals:
        print(f"  L{r:3d}: {' | '.join(vals)}")
