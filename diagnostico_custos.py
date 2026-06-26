#!/usr/bin/env python3
"""
Diagnóstico completo: analisa a planilha de custos e compara com o banco,
identificando todos os produtos que têm custo na planilha mas não foram preenchidos.
"""
import json, re, mysql.connector
import openpyxl

# ─── Conexão banco ────────────────────────────────────────────────────────────
with open('/home/ubuntu/cadastro-produtos-alfalux/.project-config.json') as f:
    config = json.load(f)
db_url = config['env_vars']['DATABASE_URL']
m = re.match(r'mysql://([^:]+):([^@]+)@([^:]+):(\d+)/([^?]+)', db_url)
user, password, host, port, dbname = m.groups()
conn = mysql.connector.connect(host=host, port=int(port), user=user, password=password, database=dbname, ssl_disabled=False)
cursor = conn.cursor(dictionary=True)

cursor.execute("SELECT id, sku, produto, familia, categoria FROM products ORDER BY sku")
all_products = cursor.fetchall()
conn.close()

# Indexar por SKU (upper)
db_by_sku = {p['sku'].upper().strip(): p for p in all_products if p['sku']}

# ─── Ler planilha ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook('/home/ubuntu/upload/TabelavendasLuminárias25.06.26final.xlsx', data_only=True)
ws = wb.active

print("=== ESTRUTURA DA PLANILHA ===")
print(f"Dimensões: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print()

# Mostrar primeiras linhas para entender a estrutura
print("=== PRIMEIRAS 10 LINHAS ===")
for row_idx in range(1, 11):
    row_vals = []
    for col_idx in range(1, 10):
        val = ws.cell(row=row_idx, column=col_idx).value
        row_vals.append(repr(val)[:30])
    print(f"  Linha {row_idx}: {row_vals}")

print()
print("=== ANÁLISE DE TODAS AS LINHAS COM DADOS ===")

# Coletar todas as linhas com SKU e custo
planilha_rows = []
for row_idx in range(1, ws.max_row + 1):
    col_a = ws.cell(row=row_idx, column=1).value  # Descrição/SKU
    col_b = ws.cell(row=row_idx, column=2).value  # Custo
    col_c = ws.cell(row=row_idx, column=3).value  # MKP Padrão
    col_d = ws.cell(row=row_idx, column=4).value  # MKP Mínimo
    
    if col_a is None:
        continue
    
    sku_raw = str(col_a).strip()
    
    # Tentar extrair custo
    custo = None
    if col_b is not None:
        try:
            custo = float(col_b)
        except:
            pass
    
    if custo and custo > 0:
        planilha_rows.append({
            'row': row_idx,
            'sku_raw': sku_raw,
            'custo': custo,
            'mkp_padrao': col_c,
            'mkp_minimo': col_d,
        })

print(f"Total de linhas com SKU e custo > 0: {len(planilha_rows)}")
print()

# ─── Tentar mapear cada SKU da planilha para o banco ─────────────────────────
# Prefixos equivalentes: LLE↔ALE, LLS↔ALS, LLP↔ALP, LDE↔ADE, LDS↔ADS, LDP↔ADP, etc.
PREFIX_MAP = {
    'LLE': ['ALE', 'LLE'],
    'LLS': ['ALS', 'LLS'],
    'LLP': ['ALP', 'LLP'],
    'LDE': ['ADE', 'LDE'],
    'LDS': ['ADS', 'LDS'],
    'LDP': ['ADP', 'LDP'],
    'ALE': ['ALE', 'LLE'],
    'ALS': ['ALS', 'LLS'],
    'ALP': ['ALP', 'LLP'],
    'ADE': ['ADE', 'LDE'],
    'ADS': ['ADS', 'LDS'],
    'ADP': ['ADP', 'LDP'],
    'LDA': ['LDA', 'ADA'],
    'ADA': ['ADA', 'LDA'],
}

def normalize_sku(sku):
    """Normaliza SKU: uppercase, sem espaços."""
    return sku.upper().strip()

def try_match(sku_raw):
    """Tenta encontrar o produto no banco para um SKU da planilha.
    Retorna (produto_banco, metodo_match) ou (None, None).
    """
    sku = normalize_sku(sku_raw)
    
    # 1. Match direto
    if sku in db_by_sku:
        return db_by_sku[sku], 'direto'
    
    # 2. Troca de prefixo
    prefix = sku[:3]
    suffix = sku[3:]
    if prefix in PREFIX_MAP:
        for alt_prefix in PREFIX_MAP[prefix]:
            alt_sku = alt_prefix + suffix
            if alt_sku in db_by_sku:
                return db_by_sku[alt_sku], f'prefixo {prefix}→{alt_prefix}'
    
    # 3. D1/D2 → D1 + D2 (substituir "/" por " + ")
    if '/' in sku:
        sku_alt = sku.replace('/', ' + ')
        if sku_alt in db_by_sku:
            return db_by_sku[sku_alt], 'D1/D2→D1+D2'
        # Também tentar com troca de prefixo
        prefix2 = sku_alt[:3]
        suffix2 = sku_alt[3:]
        if prefix2 in PREFIX_MAP:
            for alt_prefix in PREFIX_MAP[prefix2]:
                alt_sku2 = alt_prefix + suffix2
                if alt_sku2 in db_by_sku:
                    return db_by_sku[alt_sku2], f'D1/D2+prefixo {prefix2}→{alt_prefix}'
    
    return None, None

# Classificar resultados
encontrados = []
nao_encontrados = []

for row in planilha_rows:
    produto, metodo = try_match(row['sku_raw'])
    if produto:
        encontrados.append({**row, 'produto_banco': produto, 'metodo': metodo})
    else:
        nao_encontrados.append(row)

print(f"Encontrados no banco: {len(encontrados)}")
print(f"NÃO encontrados no banco: {len(nao_encontrados)}")
print()

# Mostrar os não encontrados para diagnóstico
print("=== SKUs DA PLANILHA NÃO ENCONTRADOS NO BANCO ===")
for r in nao_encontrados:
    print(f"  Linha {r['row']:4d}: {r['sku_raw']:<35} custo={r['custo']:.4f}")

print()
print("=== FAMÍLIAS DOS NÃO ENCONTRADOS ===")
# Tentar identificar famílias
familias_nao_enc = {}
for r in nao_encontrados:
    sku = r['sku_raw'].upper()
    # Extrair família do SKU (parte entre hífens)
    parts = sku.split('-')
    if len(parts) >= 2:
        familia = parts[0] + '-' + parts[1].split('.')[0]
    else:
        familia = parts[0]
    familias_nao_enc[familia] = familias_nao_enc.get(familia, 0) + 1

for fam, cnt in sorted(familias_nao_enc.items()):
    print(f"  {fam}: {cnt} produto(s)")

print()
print("=== VERIFICANDO FAMÍLIA MOON NO BANCO ===")
moon_db = [p for p in all_products if 'MOON' in (p['familia'] or '').upper() or 'MOON' in (p['sku'] or '').upper() or 'MOON' in (p['produto'] or '').upper()]
print(f"Produtos com MOON no banco: {len(moon_db)}")
for p in moon_db[:10]:
    print(f"  SKU: {p['sku']:<30} Família: {p['familia']}")

print()
print("=== VERIFICANDO MOON NA PLANILHA ===")
moon_plan = [r for r in planilha_rows if 'MOON' in r['sku_raw'].upper()]
print(f"Linhas com MOON na planilha: {len(moon_plan)}")
for r in moon_plan[:10]:
    print(f"  Linha {r['row']:4d}: {r['sku_raw']:<35} custo={r['custo']:.4f}")
