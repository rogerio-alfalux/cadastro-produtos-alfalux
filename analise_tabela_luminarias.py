#!/usr/bin/env python3
"""Analisa a planilha de custos de luminárias para entender sua estrutura."""
import openpyxl

wb = openpyxl.load_workbook('/home/ubuntu/upload/TabelavendasLuminárias25.06.26final.xlsx', data_only=True, read_only=True)
print('Abas:', wb.sheetnames)

for sh in wb.sheetnames:
    ws = wb[sh]
    print(f'\n=== Aba: {sh} ===')
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        print(row)
        count += 1
    print(f'(mostrando {count} linhas)')

wb.close()
