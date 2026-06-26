#!/usr/bin/env python3
"""
Gera planilha Excel com todos os produtos cadastrados que não têm custo preenchido.
Um produto é considerado "sem custo" quando todos os campos custoCorpo* são nulos/zero.
"""
import json, re, mysql.connector
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

with open('/home/ubuntu/cadastro-produtos-alfalux/.project-config.json') as f:
    config = json.load(f)
db_url = config['env_vars']['DATABASE_URL']
m = re.match(r'mysql://([^:]+):([^@]+)@([^:]+):(\d+)/([^?]+)', db_url)
user, password, host, port, dbname = m.groups()
conn = mysql.connector.connect(host=host, port=int(port), user=user, password=password, database=dbname, ssl_disabled=False)
cursor = conn.cursor(dictionary=True)

# Buscar todos os produtos com todos os campos de custo
cursor.execute("""
    SELECT 
        id, sku, produto, categoria, instalacao, familia,
        driverOnoff220, driverOnoffBivolt, driverDim110v, driverDimDali,
        driverDimTriac110v, driverDimTriac220v,
        driverOnoffBivoltNaoAplicavel, driverDim110vNaoAplicavel,
        driverDimDaliNaoAplicavel, driverDimTriac110vNaoAplicavel,
        driverDimTriac220vNaoAplicavel,
        custoCorpoOnoff220v, custoCorpoOnoffBivolt,
        custoCorpoDim110v, custoCorpoDimDali,
        custoCorpoDimTriac110v, custoCorpoDimTriac220v,
        custoLuminaria
    FROM products
    ORDER BY categoria, familia, sku
""")
all_products = cursor.fetchall()
conn.close()

def has_custo(p):
    """Retorna True se o produto tem pelo menos um campo de custo preenchido."""
    custo_fields = [
        'custoCorpoOnoff220v', 'custoCorpoOnoffBivolt',
        'custoCorpoDim110v', 'custoCorpoDimDali',
        'custoCorpoDimTriac110v', 'custoCorpoDimTriac220v',
        'custoLuminaria',
    ]
    for f in custo_fields:
        val = p.get(f)
        if val is not None and float(val) > 0:
            return True
    return False

def get_drivers(p):
    """Retorna string com os tipos de driver disponíveis."""
    drivers = []
    if p.get('driverOnoff220'):
        drivers.append('ON/OFF 220V')
    if p.get('driverOnoffBivolt') and not p.get('driverOnoffBivoltNaoAplicavel'):
        drivers.append('ON/OFF BIVOLT')
    if p.get('driverDim110v') and not p.get('driverDim110vNaoAplicavel'):
        drivers.append('DIM 1-10V')
    if p.get('driverDimDali') and not p.get('driverDimDaliNaoAplicavel'):
        drivers.append('DIM DALI')
    if p.get('driverDimTriac110v') and not p.get('driverDimTriac110vNaoAplicavel'):
        drivers.append('DIM TRIAC 110V')
    if p.get('driverDimTriac220v') and not p.get('driverDimTriac220vNaoAplicavel'):
        drivers.append('DIM TRIAC 220V')
    return ', '.join(drivers) if drivers else '—'

# Filtrar sem custo
sem_custo = [p for p in all_products if not has_custo(p)]
print(f"Total de produtos: {len(all_products)}")
print(f"Produtos SEM custo: {len(sem_custo)}")

# ─── Criar Excel ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sem Custo"

# Cores
COR_HEADER_BG   = "1A1A2E"   # azul escuro
COR_HEADER_FG   = "FFFFFF"
COR_CAT_BG      = "16213E"   # azul médio
COR_CAT_FG      = "E94560"   # vermelho/rosa
COR_ROW_PAR     = "0F3460"   # azul escuro suave
COR_ROW_IMPAR   = "1A1A2E"   # azul muito escuro
COR_ROW_FG      = "E0E0E0"
COR_SKU_FG      = "4FC3F7"   # azul claro
COR_ACCENT      = "E94560"

def make_border(style='thin'):
    s = Side(style=style, color="2A2A4A")
    return Border(left=s, right=s, top=s, bottom=s)

def header_style():
    return Font(name='Calibri', bold=True, color=COR_HEADER_FG, size=10)

def cat_style():
    return Font(name='Calibri', bold=True, color=COR_CAT_FG, size=10)

# ─── Cabeçalho principal ─────────────────────────────────────────────────────
ws.merge_cells('A1:G1')
title_cell = ws['A1']
title_cell.value = f"PRODUTOS SEM CUSTO CADASTRADO — {len(sem_custo)} produtos"
title_cell.font = Font(name='Calibri', bold=True, color="E94560", size=13)
title_cell.fill = PatternFill("solid", fgColor="0D0D1A")
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28

# ─── Subtítulo ───────────────────────────────────────────────────────────────
ws.merge_cells('A2:G2')
sub_cell = ws['A2']
sub_cell.value = "Produtos cadastrados no sistema que ainda não possuem custo do corpo preenchido"
sub_cell.font = Font(name='Calibri', italic=True, color="888888", size=9)
sub_cell.fill = PatternFill("solid", fgColor="0D0D1A")
sub_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 16

# ─── Cabeçalho das colunas ───────────────────────────────────────────────────
headers = ["#", "SKU", "PRODUTO", "CATEGORIA", "FAMÍLIA", "INSTALAÇÃO", "DRIVERS DISPONÍVEIS"]
col_widths = [5, 24, 45, 18, 22, 18, 40]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = Font(name='Calibri', bold=True, color=COR_HEADER_FG, size=9)
    cell.fill = PatternFill("solid", fgColor=COR_HEADER_BG)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    cell.border = make_border()
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[3].height = 20

# ─── Dados agrupados por categoria ───────────────────────────────────────────
from itertools import groupby

row_num = 4
counter = 1

# Agrupar por categoria
categorias = {}
for p in sem_custo:
    cat = p['categoria'] or 'SEM CATEGORIA'
    categorias.setdefault(cat, []).append(p)

for cat_name in sorted(categorias.keys()):
    produtos_cat = categorias[cat_name]

    # Linha de categoria
    ws.merge_cells(f'A{row_num}:G{row_num}')
    cat_cell = ws.cell(row=row_num, column=1)
    cat_cell.value = f"  {cat_name.upper()}  ({len(produtos_cat)} produto{'s' if len(produtos_cat) != 1 else ''})"
    cat_cell.font = Font(name='Calibri', bold=True, color=COR_CAT_FG, size=9)
    cat_cell.fill = PatternFill("solid", fgColor="0D1B2A")
    cat_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    cat_cell.border = make_border()
    ws.row_dimensions[row_num].height = 18
    row_num += 1

    for i, p in enumerate(produtos_cat):
        bg = COR_ROW_IMPAR if i % 2 == 0 else COR_ROW_PAR
        row_data = [
            counter,
            p['sku'] or '—',
            p['produto'] or '—',
            p['categoria'] or '—',
            p['familia'] or '—',
            p['instalacao'] or '—',
            get_drivers(p),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = make_border()
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            ws.row_dimensions[row_num].height = 16

            # Estilos específicos por coluna
            if col_idx == 1:  # #
                cell.font = Font(name='Calibri', color="666688", size=8)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 2:  # SKU
                cell.font = Font(name='Calibri', bold=True, color=COR_SKU_FG, size=9)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif col_idx == 3:  # Produto
                cell.font = Font(name='Calibri', color=COR_ROW_FG, size=9)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif col_idx == 7:  # Drivers
                cell.font = Font(name='Calibri', color="A0C4FF", size=8, italic=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.font = Font(name='Calibri', color="AAAACC", size=9)
                cell.alignment = Alignment(horizontal='left', vertical='center')

        counter += 1
        row_num += 1

# ─── Linha de total ───────────────────────────────────────────────────────────
ws.merge_cells(f'A{row_num}:G{row_num}')
total_cell = ws.cell(row=row_num, column=1)
total_cell.value = f"TOTAL: {len(sem_custo)} produtos sem custo cadastrado"
total_cell.font = Font(name='Calibri', bold=True, color="E94560", size=9)
total_cell.fill = PatternFill("solid", fgColor="0D0D1A")
total_cell.alignment = Alignment(horizontal='center', vertical='center')
total_cell.border = make_border()
ws.row_dimensions[row_num].height = 18

# ─── Segunda aba: resumo por categoria ───────────────────────────────────────
ws2 = wb.create_sheet("Resumo por Categoria")

ws2.merge_cells('A1:C1')
ws2['A1'].value = "RESUMO — PRODUTOS SEM CUSTO POR CATEGORIA"
ws2['A1'].font = Font(name='Calibri', bold=True, color="E94560", size=12)
ws2['A1'].fill = PatternFill("solid", fgColor="0D0D1A")
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 24

for col_idx, (h, w) in enumerate(zip(["CATEGORIA", "QTD SEM CUSTO", "% DO TOTAL"], [28, 18, 18]), 1):
    cell = ws2.cell(row=2, column=col_idx, value=h)
    cell.font = Font(name='Calibri', bold=True, color=COR_HEADER_FG, size=9)
    cell.fill = PatternFill("solid", fgColor=COR_HEADER_BG)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = make_border()
    ws2.column_dimensions[get_column_letter(col_idx)].width = w
ws2.row_dimensions[2].height = 18

# Contar total por categoria (incluindo com custo) para calcular %
total_por_cat = {}
for p in all_products:
    cat = p['categoria'] or 'SEM CATEGORIA'
    total_por_cat[cat] = total_por_cat.get(cat, 0) + 1

r = 3
for i, (cat_name, produtos_cat) in enumerate(sorted(categorias.items())):
    bg = COR_ROW_IMPAR if i % 2 == 0 else COR_ROW_PAR
    total_cat = total_por_cat.get(cat_name, 0)
    pct = f"{len(produtos_cat)/total_cat*100:.1f}%" if total_cat > 0 else "—"

    for col_idx, val in enumerate([cat_name, len(produtos_cat), pct], 1):
        cell = ws2.cell(row=r, column=col_idx, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = make_border()
        cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left', vertical='center')
        cell.font = Font(name='Calibri', color=COR_CAT_FG if col_idx == 1 else COR_ROW_FG, size=9,
                        bold=(col_idx == 1))
    ws2.row_dimensions[r].height = 16
    r += 1

# Total
for col_idx, val in enumerate(["TOTAL", len(sem_custo), f"{len(sem_custo)/len(all_products)*100:.1f}%"], 1):
    cell = ws2.cell(row=r, column=col_idx, value=val)
    cell.font = Font(name='Calibri', bold=True, color="E94560", size=9)
    cell.fill = PatternFill("solid", fgColor="0D0D1A")
    cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left', vertical='center')
    cell.border = make_border()
ws2.row_dimensions[r].height = 18

# Congelar painéis
ws.freeze_panes = 'A4'
ws2.freeze_panes = 'A3'

# Salvar
output_path = '/home/ubuntu/produtos_sem_custo.xlsx'
wb.save(output_path)
print(f"\nPlanilha salva em: {output_path}")
print(f"\nResumo por categoria:")
for cat_name, produtos_cat in sorted(categorias.items()):
    total_cat = total_por_cat.get(cat_name, 0)
    print(f"  {cat_name}: {len(produtos_cat)} sem custo / {total_cat} total ({len(produtos_cat)/total_cat*100:.1f}%)")
